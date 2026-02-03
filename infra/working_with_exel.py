import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict


def normalize_columns_pandas(df):
    """Normalize column names: lowercase + replace spaces with underscores."""
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    return df


def get_column_pandas(df, key, required=True):
    """
    Return the first column name in df.columns that matches expected variants.
    If required=False, returns None instead of raising error when missing.
    """
    variants = COLUMN_MAP[key]
    for variant in variants:
        for col in df.columns:
            if variant in col:
                return col
    if required:
        raise ValueError(f"Missing required column for key '{key}': variants {variants}")
    return None


def get_bug_to_tests_map(excel_path):
    """
    Map bug IDs to test case IDs from an STD Excel.
    Handles multiple bug IDs in one cell.
    """
    df = pd.read_excel(excel_path)
    df = normalize_columns_pandas(df)

    bug_col = get_column_pandas(df, "bug")
    id_col = get_column_pandas(df, "id") if "id" in df.columns else None
    if id_col is None:
        raise ValueError("No ID column found. Make sure your file has an ID column.")

    with_bugs = df[df[bug_col].apply(lambda x: isinstance(x, (int, float, str)) and not pd.isna(x))]

    bug_to_tests = defaultdict(list)
    for _, row in with_bugs.iterrows():
        raw_bug_val = str(row[bug_col])
        test_id = row[id_col]

        bug_ids = [bug.strip() for bug in str(raw_bug_val).split(",") if bug.strip()]

        for bug_id_str in bug_ids:
            # Ensure we normalize float-y values like '1234.0' → '1234'
            if bug_id_str.replace(".", "", 1).isdigit():
                if bug_id_str.endswith(".0"):
                    bug_id_str = bug_id_str[:-2]
            bug_to_tests[bug_id_str].append(test_id)

    return dict(bug_to_tests)


from openpyxl import load_workbook
from collections import defaultdict

from utils.constants import COLUMN_MAP, STDConstants, ExcelRules, VALID_TEST_RESULTS


def normalize_columns(cols):
    """Normalize column names: lowercase + replace spaces with underscores."""
    return [str(col).strip().lower().replace(" ", "_") if col is not None else "" for col in cols]



def get_column(headers, key, required=True):
    """Return first header that matches any variant of key. If required=False, returns None when missing."""
    variants = COLUMN_MAP.get(key, [])
    if not variants:
        return None if not required else _raise_missing(key)
    for variant in variants:
        for h in headers:
            if variant in h:
                return h
    if required:
        raise ValueError(f"Missing required column for key '{key}'")
    return None


def _raise_missing(key):
    raise ValueError(f"Missing required column for key '{key}'")


def is_precondition_row(row, headers):
    """Check if row is a precondition row (only id/headline/desc filled, rest empty)."""
    possible_must_have = ExcelRules.POSSIBLE_MUST_HAVE_COLUMNS
    must_have = [h for h in headers if h in possible_must_have]
    other_cols = [h for h in headers if h not in must_have]

    for col in must_have:
        val = row.get(col)
        if val is None or str(val).strip() == '':
            return False

    for col in other_cols:
        val = row.get(col)
        if val is not None and str(val).strip() != '':
            return False

    return True


def is_precondition_expected_na(row, expected_col):
    """True when Expected Result is N/A (precondition test - Results/Actual/Bug must be empty)."""
    exp = row.get(expected_col)
    exp_str = str(exp or "").strip().lower()
    return exp_str == STDConstants.N_A


def validate_and_summarize(file_path):
    """Validate STD rules using openpyxl."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    headers = normalize_columns([cell.value for cell in ws[1]])
    expected_col = get_column(headers, "expected")
    results_col = get_column(headers, "results")
    bug_col = get_column(headers, "bug")
    comment_col = get_column(headers, "comment", required=False)
    id_col = get_column(headers, "id")
    actual_col = get_column(headers, "actual", required=False)

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        # Strip strings
        for k, v in row_dict.items():
            if isinstance(v, str):
                row_dict[k] = v.strip()
        data.append(row_dict)

    # Masks for rules
    rule1_rows = []
    rule2_rows = []
    rule3_rows = []
    rule4_rows = []
    rule5_rows = []
    rule6_rows = []
    rule7_rows = []
    rule8_rows = []

    for row in data:
        res = str(row.get(results_col) or '').strip()
        res_lower = res.lower()
        exp = row.get(expected_col)
        exp_lower = str(exp or "").strip().lower()
        bug = row.get(bug_col)
        bug_empty = bug is None or str(bug).strip() == ""

        # Rule1: expected filled AND results empty; exclude precondition (Expected=N/A)
        if (exp and exp_lower != STDConstants.N_A and
                (res == '' or res_lower in [STDConstants.NONE, STDConstants.NAN]) and
                res_lower not in [STDConstants.N_A, STDConstants.NOT_TESTED]):
            rule1_rows.append(row)

        # Rule2: results filled AND expected empty, excluding precondition
        if (res != '' and exp in [None, '']) and not is_precondition_row(row, headers):
            rule2_rows.append(row)

        # Rule3: bug filled AND results = pass
        if bug and res_lower == STDConstants.PASS:
            rule3_rows.append(row)

        # Rule4: bug empty AND results = fail
        if bug_empty and res_lower == STDConstants.FAIL:
            rule4_rows.append(row)

        # Rule5: Actual Results vs Test Result (only when Actual column exists)
        if actual_col:
            actual_val = str(row.get(actual_col) or '').strip()
            actual_lower = actual_val.lower()
            test_result_val = res_lower

            if test_result_val == STDConstants.PASS:
                if actual_lower != STDConstants.ACTUAL_PASS_VALUE.lower():
                    rule5_rows.append(row)
            elif test_result_val == STDConstants.FAIL:
                if not actual_val.upper().startswith(STDConstants.ACTUAL_FAIL_PREFIX.upper()):
                    rule5_rows.append(row)
                else:
                    after_comma = (actual_val.split(",", 1)[1].strip() if "," in actual_val else "").strip()
                    if not after_comma:
                        rule5_rows.append(row)
            elif test_result_val == STDConstants.NOT_TESTED or test_result_val == STDConstants.N_A:
                if actual_lower != STDConstants.N_A:
                    rule5_rows.append(row)

        # Rule6: Precondition (Expected=N/A) must have empty Results, Actual, Bug
        if is_precondition_expected_na(row, expected_col):
            has_results = res != ""
            has_actual = actual_col and str(row.get(actual_col) or "").strip() != ""
            if has_results or has_actual or not bug_empty:
                rule6_rows.append(row)

        # Rule7: Test Result = N/A must have a comment (when Comment column exists)
        if comment_col and res_lower == STDConstants.N_A:
            comment_val = str(row.get(comment_col) or "").strip()
            if not comment_val:
                rule7_rows.append(row)

        # Rule8: Test Results must be one of Pass, Fail, Not Tested, N/A (when not empty)
        if res != "" and res_lower not in VALID_TEST_RESULTS:
            rule8_rows.append(row)

    rules = {
        "Rule1": rule1_rows,
        "Rule2": rule2_rows,
        "Rule3": rule3_rows,
        "Rule4": rule4_rows,
        "Rule5": rule5_rows if actual_col else [],
        "Rule6": rule6_rows,
        "Rule7": rule7_rows if comment_col else [],
        "Rule8": rule8_rows,
    }

    # Print summary
    for rule_name, rows in rules.items():
        if not rows:
            print(f"{rule_name}: PASS — no violations found.")
        else:
            print(f"{rule_name}: FAIL — {len(rows)} violation(s) found:")
            for r in rows[:10]:
                print(f"  {r.get(id_col)}  {r.get('headline', '')}")
        print("-" * 40)

    return rules
